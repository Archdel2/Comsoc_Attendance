"""
Main application for the Attendance Management System
"""

import sys
import os
import qrcode
from PyQt5.QtWidgets import QApplication, QMainWindow, QStackedWidget, QMessageBox, QComboBox, QTableWidgetItem, QFileDialog
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont
import openpyxl
from datetime import datetime

from database import DatabaseManager
from ui_pages import MainPage, EventsPage, RecordsPage, MasterlistPage, AttendancePage, ScannerPage
from camera_scanner import CameraScanner
from config import APP_TITLE, APP_WIDTH, APP_HEIGHT, ATTENDANCE_STATUSES


class MainWindow(QMainWindow):
    """Main application window"""
    
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.current_event_id = None
        self.current_record_id = None
        self.setup_ui()
        self.setup_camera()
        self.initialize_data()
    
    def setup_ui(self):
        """Setup the main window UI"""
        self.setWindowTitle(APP_TITLE)
        self.setGeometry(100, 100, APP_WIDTH, APP_HEIGHT)
        
        self.central_widget = QStackedWidget()
        self.setCentralWidget(self.central_widget)
        
        self.main_page = MainPage(self)
        self.events_page = EventsPage(self)
        self.records_page = RecordsPage(self)
        self.masterlist_page = MasterlistPage(self)
        self.attendance_page = AttendancePage(self)
        self.scanner_page = ScannerPage(self)
        
        self.central_widget.addWidget(self.main_page)
        self.central_widget.addWidget(self.events_page)
        self.central_widget.addWidget(self.records_page)
        self.central_widget.addWidget(self.masterlist_page)
        self.central_widget.addWidget(self.attendance_page)
        self.central_widget.addWidget(self.scanner_page)
    
    def setup_camera(self):
        """Setup camera scanner for the scanner page"""
        self.camera_scanner = CameraScanner(
            self.scanner_page.camera_label,
            self.scanner_page.scanner_status,
            self
        )
    
    def initialize_data(self):
        """Initialize sample data if no students exist"""
        if not self.db.get_all_students():
            self.db.import_sample_data()
    
    def show_events_page(self):
        """Show the events page and populate the table"""
        self.populate_events_table()
        self.central_widget.setCurrentWidget(self.events_page)
    
    def show_masterlist_page(self):
        """Show the masterlist page and populate the table"""
        self.populate_masterlist_table()
        self.central_widget.setCurrentWidget(self.masterlist_page)
    
    def show_scanner_page(self):
        """Show the scanner page and start camera"""
        self.central_widget.setCurrentWidget(self.scanner_page)
        self.camera_scanner.start_camera()
    
    def stop_camera_and_go_back(self):
        """Stop camera and return to attendance page"""
        self.camera_scanner.stop_camera()
        self.central_widget.setCurrentWidget(self.attendance_page)
    
    def populate_events_table(self):
        """Populate the events table with data from database"""
        events = self.db.get_all_events()
        self.events_page.events_table.setRowCount(len(events))
        
        for row, event in enumerate(events):
            event_name = str(event['event_name'])
            event_date = str(event['event_date'])
            
            self.events_page.events_table.setItem(row, 0, QTableWidgetItem(event_name))
            self.events_page.events_table.setItem(row, 1, QTableWidgetItem(event_date))
    
    def populate_masterlist_table(self):
        """Populate the masterlist table with student data"""
        students = self.db.get_all_students()
        self.masterlist_page.masterlist_table.setRowCount(len(students))
        
        for row, student in enumerate(students):
            student_id = str(student['student_id'])
            fname = str(student['fname'])
            year_level = str(student['year_level'])
            course = str(student['course'])
            
            self.masterlist_page.masterlist_table.setItem(row, 0, QTableWidgetItem(student_id))
            self.masterlist_page.masterlist_table.setItem(row, 1, QTableWidgetItem(fname))
            self.masterlist_page.masterlist_table.setItem(row, 2, QTableWidgetItem(year_level))
            self.masterlist_page.masterlist_table.setItem(row, 3, QTableWidgetItem(course))
        
        self.filter_masterlist_table()
    
    def populate_attendance_table(self, event_id):
        """Populate the attendance table for a specific event"""
        attendance = self.db.get_attendance_for_event(event_id)
        self.attendance_page.attendance_table.setRowCount(len(attendance))
        
        for row, record in enumerate(attendance):
            student_id = str(record['student_id'])
            student_fname = str(record['student_fname'])
            student_year_level = str(record['student_year_level'])
            student_course = str(record['student_course'])
            status = str(record['status'])
            timestamp = str(record['timestamp']) if record['timestamp'] else ''
            
            self.attendance_page.attendance_table.setItem(row, 0, QTableWidgetItem(student_id))
            self.attendance_page.attendance_table.setItem(row, 1, QTableWidgetItem(student_fname))
            self.attendance_page.attendance_table.setItem(row, 2, QTableWidgetItem(student_year_level))
            self.attendance_page.attendance_table.setItem(row, 3, QTableWidgetItem(student_course))
            self.attendance_page.attendance_table.setItem(row, 4, QTableWidgetItem(timestamp))
            
            status_combo = QComboBox()
            status_combo.addItems(ATTENDANCE_STATUSES)
            status_combo.setCurrentText(status)
            status_combo.currentTextChanged.connect(
                lambda status, row=row, event_id=event_id, student_id=student_id: 
                self.db.update_attendance_status(event_id, student_id, status)
            )
            self.attendance_page.attendance_table.setCellWidget(row, 5, status_combo)
        
        self.filter_attendance_table()
    
    def populate_records_table(self, event_id):
        """Populate the records table for a specific event"""
        records = self.db.get_records_for_event(event_id)
        self.records_page.records_table.setRowCount(len(records))
        
        for row, record in enumerate(records):
            record_name = str(record['record_name'])
            created_at = str(record['created_at'])
            
            self.records_page.records_table.setItem(row, 0, QTableWidgetItem(record_name))
            self.records_page.records_table.setItem(row, 1, QTableWidgetItem(created_at))
    
    def populate_students_table(self, record_id):
        """Populate the students table for a specific record"""
        students = self.db.get_students_for_record(record_id)
        self.attendance_page.attendance_table.setRowCount(len(students))
        
        for row, student in enumerate(students):
            student_id = str(student['student_id'])
            student_fname = str(student['student_fname'])
            student_year_level = str(student['student_year_level'])
            student_course = str(student['student_course'])
            status = str(student['status'])
            timestamp = str(student['timestamp']) if student['timestamp'] else ''
            
            self.attendance_page.attendance_table.setItem(row, 0, QTableWidgetItem(student_id))
            self.attendance_page.attendance_table.setItem(row, 1, QTableWidgetItem(student_fname))
            self.attendance_page.attendance_table.setItem(row, 2, QTableWidgetItem(student_year_level))
            self.attendance_page.attendance_table.setItem(row, 3, QTableWidgetItem(student_course))
            self.attendance_page.attendance_table.setItem(row, 4, QTableWidgetItem(timestamp))
            
            status_combo = QComboBox()
            status_combo.addItems(ATTENDANCE_STATUSES)
            status_combo.setCurrentText(status)
            status_combo.currentTextChanged.connect(
                lambda status, row=row, record_id=record_id, student_id=student_id: 
                self.update_attendance_status_for_record(record_id, student_id, status)
            )
            self.attendance_page.attendance_table.setCellWidget(row, 5, status_combo)
        
        self.filter_attendance_table()
    
    def update_attendance_status_for_record(self, record_id, student_id, status):
        """Update attendance status for a specific student in a specific record"""
        try:
            with self.db.conn.cursor() as cursor:
                cursor.execute('''UPDATE Attendance 
                                  SET status = %s, timestamp = NOW()
                                  WHERE record_id = %s AND student_id = %s''',
                               (status, record_id, student_id))
                self.db.conn.commit()
                return True
        except Exception as err:
            print(f"Error updating attendance status: {err}")
            return False
    
    def filter_attendance_table(self):
        """Filter the attendance table based on search text and status filter"""
        search_text = self.attendance_page.search_input.text().lower()
        status_filter = self.attendance_page.status_filter.currentText()
        
        if hasattr(self, 'current_record_id') and self.current_record_id:
            all_students = self.db.get_students_for_record(self.current_record_id)
            is_record_context = True
        elif hasattr(self, 'current_event_id') and self.current_event_id:
            all_students = self.db.get_attendance_for_event(self.current_event_id)
            is_record_context = False
        else:
            return

        filtered_students = []
        for student in all_students:
            matches_search = (
                search_text in str(student['student_id']).lower() or
                search_text in str(student['student_fname']).lower() or
                search_text in str(student['student_year_level']).lower() or
                search_text in str(student['student_course']).lower()
            )
            
            matches_status = (status_filter == "All Statuses" or 
                            student['status'] == status_filter)
            
            if matches_search and matches_status:
                filtered_students.append(student)
        
        self.attendance_page.attendance_table.setRowCount(len(filtered_students))
        
        for row, student in enumerate(filtered_students):
            student_id = str(student['student_id'])
            student_fname = str(student['student_fname'])
            student_year_level = str(student['student_year_level'])
            student_course = str(student['student_course'])
            status = str(student['status'])
            timestamp = str(student['timestamp']) if student['timestamp'] else ''
            
            self.attendance_page.attendance_table.setItem(row, 0, QTableWidgetItem(student_id))
            self.attendance_page.attendance_table.setItem(row, 1, QTableWidgetItem(student_fname))
            self.attendance_page.attendance_table.setItem(row, 2, QTableWidgetItem(student_year_level))
            self.attendance_page.attendance_table.setItem(row, 3, QTableWidgetItem(student_course))
            self.attendance_page.attendance_table.setItem(row, 4, QTableWidgetItem(timestamp))
            
            status_combo = QComboBox()
            status_combo.addItems(ATTENDANCE_STATUSES)
            status_combo.setCurrentText(status)
            
            if is_record_context:
                status_combo.currentTextChanged.connect(
                    lambda status, row=row, record_id=self.current_record_id, student_id=student_id: 
                    self.update_attendance_status_for_record(record_id, student_id, status)
                )
            else:
                status_combo.currentTextChanged.connect(
                    lambda status, row=row, event_id=self.current_event_id, student_id=student_id: 
                    self.db.update_attendance_status(event_id, student_id, status)
                )
            
            self.attendance_page.attendance_table.setCellWidget(row, 5, status_combo)
    
    def filter_masterlist_table(self):
        """Filter the masterlist table based on search text"""
        search_text = self.masterlist_page.masterlist_search_input.text().lower()
        
        all_students = self.db.get_all_students()
        
        filtered_students = []
        for student in all_students:
            matches_search = (
                search_text in str(student['student_id']).lower() or
                search_text in str(student['fname']).lower() or
                search_text in str(student['year_level']).lower() or
                search_text in str(student['course']).lower()
            )
            
            if matches_search:
                filtered_students.append(student)
        
        self.masterlist_page.masterlist_table.setRowCount(len(filtered_students))
        
        for row, student in enumerate(filtered_students):
            student_id = str(student['student_id'])
            fname = str(student['fname'])
            year_level = str(student['year_level'])
            course = str(student['course'])
            
            self.masterlist_page.masterlist_table.setItem(row, 0, QTableWidgetItem(student_id))
            self.masterlist_page.masterlist_table.setItem(row, 1, QTableWidgetItem(fname))
            self.masterlist_page.masterlist_table.setItem(row, 2, QTableWidgetItem(year_level))
            self.masterlist_page.masterlist_table.setItem(row, 3, QTableWidgetItem(course))
    
    def export_attendance_to_excel(self):
        """Export the current attendance table to an Excel file"""
        try:
            if hasattr(self, 'current_record_id') and self.current_record_id:
                all_students = self.db.get_students_for_record(self.current_record_id)
                context_type = "Record"
                context_name = self.attendance_page.attendance_title.text().replace("Students for Record: ", "")
            elif hasattr(self, 'current_event_id') and self.current_event_id:
                all_students = self.db.get_attendance_for_event(self.current_event_id)
                context_type = "Event"
                context_name = self.attendance_page.attendance_title.text().replace("Attendance for: ", "")
            else:
                QMessageBox.warning(self, "Error", "No attendance data to export.")
                return
            
            if not all_students:
                QMessageBox.warning(self, "Error", "No attendance data to export.")
                return
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Attendance Data"
            
            ws['A1'] = f"Attendance Report - {context_type}: {context_name}"
            ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A3'] = f"Total Students: {len(all_students)}"
            
            headers = ["Student ID", "First Name", "Year Level", "Course", "Status", "Timestamp"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=5, column=col, value=header)
            
            for row, student in enumerate(all_students, 6):
                ws.cell(row=row, column=1, value=student['student_id'])
                ws.cell(row=row, column=2, value=student['student_fname'])
                ws.cell(row=row, column=3, value=student['student_year_level'])
                ws.cell(row=row, column=4, value=student['student_course'])
                ws.cell(row=row, column=5, value=student['status'])
                ws.cell(row=row, column=6, value=str(student['timestamp']) if student['timestamp'] else '')
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "Save Attendance Report", 
                f"attendance_report_{context_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                wb.save(file_path)
                QMessageBox.information(self, "Success", f"Attendance report exported successfully to:\n{file_path}")
            else:
                QMessageBox.information(self, "Cancelled", "Export was cancelled.")
                
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export attendance report:\n{str(e)}")
            print(f"Export error: {e}")
    
    def add_event(self):
        """Add a new event to the database"""
        event_name = self.events_page.event_name_input.text().strip()
        if event_name:
            self.db.create_event(event_name)
            self.events_page.event_name_input.clear()
            self.populate_events_table()
            QMessageBox.information(self, "Success", f"Event '{event_name}' created successfully!")
        else:
            QMessageBox.warning(self, "Error", "Please enter an event name.")
    
    def add_attendance_record(self):
        """Add a new attendance record to the database"""
        record_name = self.records_page.record_name_input.text().strip()
        if record_name and self.current_event_id:
            self.db.create_attendance_record(record_name, self.current_event_id)
            self.records_page.record_name_input.clear()
            self.populate_records_table(self.current_event_id)
            QMessageBox.information(self, "Success", f"Record '{record_name}' created successfully!")
        else:
            QMessageBox.warning(self, "Error", "Please enter a record name.")
    
    def view_event_attendance(self, row):
        """View attendance for a specific event"""
        event_id = int(self.events_page.events_table.item(row, 0).text())
        event_name = self.events_page.events_table.item(row, 1).text()
        
        self.current_event_id = event_id
        self.attendance_page.attendance_title.setText(f"Attendance for: {event_name}")
        self.populate_attendance_table(event_id)
        
        self.central_widget.setCurrentWidget(self.attendance_page)
    
    def view_event_records(self, row):
        """View records for a specific event"""
        event_name = self.events_page.events_table.item(row, 0).text()
        
        events = self.db.get_all_events()
        event_id = None
        for event in events:
            if event['event_name'] == event_name:
                event_id = event['event_id']
                break
        
        if event_id:
            self.current_event_id = event_id
            self.records_page.records_title.setText(f"Records for Event: {event_name}")
            self.populate_records_table(event_id)
            
            self.central_widget.setCurrentWidget(self.records_page)
        else:
            QMessageBox.warning(self, "Error", "Could not find event information.")
    
    def view_record_students(self, row):
        """View students for a specific record"""
        record_name = self.records_page.records_table.item(row, 0).text()
        
        records = self.db.get_records_for_event(self.current_event_id)
        record_id = None
        for record in records:
            if record['record_name'] == record_name:
                record_id = record['record_id']
                break
        
        if record_id:
            self.current_record_id = record_id
            self.attendance_page.attendance_title.setText(f"Students for Record: {record_name}")
            self.populate_students_table(record_id)
            
            self.central_widget.setCurrentWidget(self.attendance_page)
        else:
            QMessageBox.warning(self, "Error", "Could not find record information.")

    def generate_qr_codes_for_all_students(self):
        """Generate QR codes for all students, grouped by year level"""
        print("generate qr_codes for all students")
        students = self.db.get_all_students()
        if not students:
            QMessageBox.warning(self, "No Students", "No students found in the database.")
            return

        output_dir = os.path.join(os.getcwd(), "student_qrcodes")
        os.makedirs(output_dir, exist_ok=True)

        year_groups = {}
        for student in students:
            year = str(student['year_level'])
            if year not in year_groups:
                year_groups[year] = []
            year_groups[year].append(student)

        for year, students_in_year in year_groups.items():
            year_dir = os.path.join(output_dir, year)
            os.makedirs(year_dir, exist_ok=True)
            for student in students_in_year:
                student_id = str(student['student_id'])
                img = qrcode.make(student_id)
                img.save(os.path.join(year_dir, f"{student_id}.png"))
        
        QMessageBox.information(self, "QR Codes Generated", f"QR codes saved in:\n{output_dir}")
    
    def closeEvent(self, event):
        """Handle application close event"""
        self.camera_scanner.stop_camera()
        self.db.close()
        event.accept()


def main():
    """Main application entry point"""
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
